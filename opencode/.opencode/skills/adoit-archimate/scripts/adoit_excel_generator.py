#!/usr/bin/env python3
"""
ADOIT ArchiMate Excel Import Generator

This tool generates Excel files compatible with ADOIT's import interface,
allowing programmatic creation of ArchiMate architecture models.

Usage:
    from adoit_excel_generator import ADOITExcelGenerator
    
    gen = ADOITExcelGenerator()
    
    # Add elements
    gen.add_capability("AI Orchestration", description="Core AI platform capability")
    gen.add_application_component("LLM Gateway", description="API gateway for LLM access")
    
    # Add relationships (via columns)
    gen.add_capability("Data Management", composition_capabilities=["Vector Store", "Knowledge Base"])
    
    # Generate Excel file
    gen.save("my_architecture.xlsx")
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from datetime import datetime
import shutil
import os


@dataclass
class ArchiMateElement:
    """Represents an ArchiMate element with attributes and relationships"""
    name: str
    element_type: str
    id: str = ""
    description: str = ""
    specialisation: str = ""
    attributes: Dict[str, Any] = field(default_factory=dict)
    relationships: Dict[str, List[str]] = field(default_factory=dict)


class ADOITExcelGenerator:
    """
    Generates ADOIT-compatible Excel import files for ArchiMate models.
    
    Key Features:
    - Supports all 61 ArchiMate element types
    - Handles relationships via named columns (->Target, <-Source)
    - Multiple targets separated by semicolons
    - Preserves template structure for compatibility
    """
    
    # ArchiMate element types supported by ADOIT
    ELEMENT_TYPES = [
        "Application Collaboration", "Application Component", "Application Event",
        "Application Function", "Application Interaction", "Application Interface",
        "Application Process", "Application Service", "Artifact", "Assessment",
        "Business Actor", "Business Collaboration", "Business Event", "Business Function",
        "Business Interaction", "Business Interface", "Business Object", "Business Process",
        "Business Role", "Business Service", "Capability", "Communication Network",
        "Constraint", "Contract", "Course of Action", "Data Object", "Deliverable",
        "Device", "Distribution Network", "Driver", "Equipment", "Facility", "Gap",
        "Goal", "Grouping", "Implementation Event", "Junction", "Location", "Material",
        "Meaning", "Node", "Outcome", "Path", "Plateau", "Principle", "Product",
        "Representation", "Requirement", "Resource", "Stakeholder", "System Software",
        "Technology Collaboration", "Technology Event", "Technology Function",
        "Technology Interaction", "Technology Interface", "Technology Process",
        "Technology Service", "Value", "Value Stream", "Work Package"
    ]
    
    # Core ArchiMate relationships
    RELATIONSHIPS = [
        "Composition", "Aggregation", "Assignment", "Realization", "Serving",
        "Access", "Influence", "Triggering", "Flow", "Specialization", "Association"
    ]
    
    def __init__(self, template_path: str = None):
        """
        Initialize the generator.
        
        Args:
            template_path: Path to ADOIT template Excel file (optional)
        """
        self.template_path = template_path
        self.elements: Dict[str, List[ArchiMateElement]] = {et: [] for et in self.ELEMENT_TYPES}
        self._template_columns: Dict[str, List[str]] = {}
        
        if template_path and os.path.exists(template_path):
            self._load_template_structure()
    
    def _load_template_structure(self):
        """Load column structure from template for each sheet"""
        sheets = pd.read_excel(self.template_path, sheet_name=None)
        for sheet_name, df in sheets.items():
            if sheet_name != "Overview":
                self._template_columns[sheet_name] = list(df.columns)
    
    def _parse_relationship_column(self, col_name: str) -> tuple:
        """
        Parse relationship column name.
        Returns (relationship_type, direction, target_type)
        
        Examples:
            "Composition (->Capability)" -> ("Composition", "->", "Capability")
            "Realization (<-Application Component)" -> ("Realization", "<-", "Application Component")
        """
        if "(->" in col_name:
            parts = col_name.split("(->")
            return (parts[0].strip(), "->", parts[1].rstrip(")").strip())
        elif "(<-" in col_name:
            parts = col_name.split("(<-")
            return (parts[0].strip(), "<-", parts[1].rstrip(")").strip())
        return (None, None, None)
    
    def add_element(
        self,
        element_type: str,
        name: str,
        id: str = "",
        description: str = "",
        specialisation: str = "",
        **kwargs
    ) -> ArchiMateElement:
        """
        Add an ArchiMate element.
        
        Args:
            element_type: Type of element (e.g., "Capability", "Application Component")
            name: Element name
            id: Optional external ID
            description: Element description
            specialisation: Specialisation value (enum)
            **kwargs: Additional attributes and relationships
                - For relationships, use format: composition_capability=["Child1", "Child2"]
                  or serving_capability="ServiceTarget"
        
        Returns:
            The created ArchiMateElement
        """
        if element_type not in self.ELEMENT_TYPES:
            raise ValueError(f"Unknown element type: {element_type}. Valid types: {self.ELEMENT_TYPES}")
        
        # Separate attributes from relationships
        attributes = {}
        relationships = {}
        
        for key, value in kwargs.items():
            # Check if this is a relationship (format: relationtype_targettype)
            rel_match = False
            for rel in self.RELATIONSHIPS:
                rel_lower = rel.lower().replace(" ", "_")
                if key.startswith(rel_lower + "_"):
                    target_type = key[len(rel_lower) + 1:].replace("_", " ").title()
                    if isinstance(value, str):
                        value = [value]
                    relationships[f"{rel} (->{target_type})"] = value
                    rel_match = True
                    break
            
            if not rel_match:
                attributes[key] = value
        
        element = ArchiMateElement(
            name=name,
            element_type=element_type,
            id=id,
            description=description,
            specialisation=specialisation,
            attributes=attributes,
            relationships=relationships
        )
        
        self.elements[element_type].append(element)
        return element
    
    # Convenience methods for common element types
    def add_capability(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Capability element"""
        return self.add_element("Capability", name, **kwargs)
    
    def add_application_component(self, name: str, **kwargs) -> ArchiMateElement:
        """Add an Application Component element"""
        return self.add_element("Application Component", name, **kwargs)
    
    def add_application_service(self, name: str, **kwargs) -> ArchiMateElement:
        """Add an Application Service element"""
        return self.add_element("Application Service", name, **kwargs)
    
    def add_application_interface(self, name: str, **kwargs) -> ArchiMateElement:
        """Add an Application Interface element"""
        return self.add_element("Application Interface", name, **kwargs)
    
    def add_application_function(self, name: str, **kwargs) -> ArchiMateElement:
        """Add an Application Function element"""
        return self.add_element("Application Function", name, **kwargs)
    
    def add_business_process(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Business Process element"""
        return self.add_element("Business Process", name, **kwargs)
    
    def add_business_actor(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Business Actor element"""
        return self.add_element("Business Actor", name, **kwargs)
    
    def add_business_role(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Business Role element"""
        return self.add_element("Business Role", name, **kwargs)
    
    def add_business_service(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Business Service element"""
        return self.add_element("Business Service", name, **kwargs)
    
    def add_data_object(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Data Object element"""
        return self.add_element("Data Object", name, **kwargs)
    
    def add_technology_service(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Technology Service element"""
        return self.add_element("Technology Service", name, **kwargs)
    
    def add_node(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Node element"""
        return self.add_element("Node", name, **kwargs)
    
    def add_device(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Device element"""
        return self.add_element("Device", name, **kwargs)
    
    def add_system_software(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a System Software element"""
        return self.add_element("System Software", name, **kwargs)
    
    def add_goal(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Goal element"""
        return self.add_element("Goal", name, **kwargs)
    
    def add_principle(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Principle element"""
        return self.add_element("Principle", name, **kwargs)
    
    def add_requirement(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Requirement element"""
        return self.add_element("Requirement", name, **kwargs)
    
    def add_constraint(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Constraint element"""
        return self.add_element("Constraint", name, **kwargs)
    
    def add_stakeholder(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Stakeholder element"""
        return self.add_element("Stakeholder", name, **kwargs)
    
    def add_driver(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Driver element"""
        return self.add_element("Driver", name, **kwargs)
    
    def add_work_package(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Work Package element"""
        return self.add_element("Work Package", name, **kwargs)
    
    def add_plateau(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Plateau element"""
        return self.add_element("Plateau", name, **kwargs)
    
    def add_gap(self, name: str, **kwargs) -> ArchiMateElement:
        """Add a Gap element"""
        return self.add_element("Gap", name, **kwargs)
    
    def _build_row_data(self, element: ArchiMateElement, columns: List[str]) -> Dict[str, Any]:
        """Build a row dictionary from an element for the given columns"""
        row = {}
        
        for col in columns:
            col_lower = col.lower()
            col_base = col.split("(")[0].strip().lower()  # Get base name without type hint
            
            # Core attributes - match exact column patterns
            if col_lower == "name (simple)" or col_base == "name":
                row[col] = element.name
            elif col_lower == "id (simple)" or col_base == "id":
                row[col] = element.id
            elif col_lower == "description (simple)" or col_base == "description":
                row[col] = element.description
            elif col_lower == "specialisation (enum)" or col_base == "specialisation":
                row[col] = element.specialisation
            elif "(->" in col or "(<-" in col:
                # Relationship columns
                rel_type, direction, target_type = self._parse_relationship_column(col)
                if rel_type and direction == "->":
                    # Outgoing relationship
                    rel_key = f"{rel_type} (->{target_type})"
                    if rel_key in element.relationships:
                        row[col] = ";".join(element.relationships[rel_key])
                    else:
                        row[col] = ""
                elif rel_type and direction == "<-":
                    # Incoming relationship - typically left empty (defined on source)
                    row[col] = ""
            else:
                # Check custom attributes
                attr_name = col_base.replace(" ", "_")
                if attr_name in element.attributes:
                    row[col] = element.attributes[attr_name]
                else:
                    row[col] = ""
        
        return row
    
    def save(self, output_path: str, use_template: bool = True):
        """
        Save the model to an Excel file.
        
        Args:
            output_path: Path for output Excel file
            use_template: If True and template exists, copy and modify template
        """
        if use_template and self.template_path and os.path.exists(self.template_path):
            # Copy template and modify
            shutil.copy(self.template_path, output_path)
            self._update_template(output_path)
        else:
            # Generate from scratch
            self._generate_new(output_path)
    
    def _update_template(self, output_path: str):
        """Update a copied template with element data"""
        wb = load_workbook(output_path)
        
        for element_type, elements in self.elements.items():
            if not elements:
                continue
            
            if element_type not in wb.sheetnames:
                continue
            
            sheet = wb[element_type]
            columns = self._template_columns.get(element_type, [])
            
            if not columns:
                # Read columns from first row
                columns = [cell.value for cell in sheet[1] if cell.value]
            
            # Add data rows starting from row 2
            for row_idx, element in enumerate(elements, start=2):
                row_data = self._build_row_data(element, columns)
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
    
    def _generate_new(self, output_path: str):
        """Generate a new Excel file without template"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Create Overview sheet
            overview_data = {
                "ADOIT Import File": ["Generated by ADOIT Excel Generator"],
                "Date": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                "Elements": [sum(len(els) for els in self.elements.values())]
            }
            pd.DataFrame(overview_data).to_excel(writer, sheet_name="Overview", index=False)
            
            # Create sheets for each element type with data
            for element_type, elements in self.elements.items():
                if not elements:
                    continue
                
                # Define columns based on element type
                columns = self._get_default_columns(element_type)
                
                # Build data rows
                rows = []
                for element in elements:
                    row = self._build_row_data(element, columns)
                    rows.append(row)
                
                df = pd.DataFrame(rows, columns=columns)
                df.to_excel(writer, sheet_name=element_type, index=False)
    
    def _get_default_columns(self, element_type: str) -> List[str]:
        """Get default columns for an element type"""
        # Basic columns common to all types
        base_columns = [
            "Name (simple)",
            "ID (simple)",
            "Description (simple)",
        ]
        
        # Add relationship columns based on element type
        if element_type == "Capability":
            return base_columns + [
                "Level (simple)",
                "Composition (->Capability)",
                "Aggregation (->Capability)",
                "Serving (->Capability)",
                "Realization (->Goal)",
                "Association (->Application Component)",
            ]
        elif element_type == "Application Component":
            return base_columns + [
                "Specialisation (enum)",
                "Composition (->Application Component)",
                "Aggregation (->Application Component)",
                "Serving (->Application Service)",
                "Realization (->Capability)",
                "Assignment (->Application Function)",
            ]
        elif element_type == "Application Service":
            return base_columns + [
                "Serving (->Business Process)",
                "Serving (->Application Service)",
                "Realization (->Application Component)",
            ]
        else:
            return base_columns
    
    def get_summary(self) -> Dict[str, int]:
        """Get summary of elements by type"""
        return {et: len(els) for et, els in self.elements.items() if els}
    
    def __repr__(self) -> str:
        summary = self.get_summary()
        total = sum(summary.values())
        return f"ADOITExcelGenerator({total} elements: {summary})"


# Example usage and demonstration
def create_sample_ai_architecture():
    """
    Create a sample AI Platform architecture model.
    Demonstrates how to use the generator.
    """
    gen = ADOITExcelGenerator()
    
    # Top-level Capabilities
    gen.add_capability(
        "AI Platform",
        description="Enterprise AI Platform for LLM-based applications",
        composition_capability=["AI Orchestration", "Data Management", "Security & Governance", "Infrastructure"]
    )
    
    # Second-level Capabilities
    gen.add_capability(
        "AI Orchestration",
        description="Orchestration of AI agents and workflows",
        composition_capability=["Agent Management", "Workflow Automation", "Model Hosting"]
    )
    
    gen.add_capability(
        "Data Management",
        description="Data storage and retrieval for AI",
        composition_capability=["Vector Store", "Knowledge Base", "Data Pipeline"]
    )
    
    gen.add_capability(
        "Security & Governance",
        description="Security and compliance for AI systems",
        composition_capability=["Access Control", "Policy Enforcement", "Audit & Compliance"]
    )
    
    gen.add_capability(
        "Infrastructure",
        description="Infrastructure services for AI workloads",
        composition_capability=["Compute", "Networking", "Observability"]
    )
    
    # Application Components
    gen.add_application_component(
        "LLM Gateway",
        description="API Gateway for LLM model access",
        realization_capability=["Model Hosting"]
    )
    
    gen.add_application_component(
        "Agent Framework",
        description="Framework for building AI agents",
        realization_capability=["Agent Management"]
    )
    
    gen.add_application_component(
        "Vector Database",
        description="Database for embeddings and semantic search",
        realization_capability=["Vector Store"]
    )
    
    gen.add_application_component(
        "Workflow Engine",
        description="Engine for AI workflow automation",
        realization_capability=["Workflow Automation"]
    )
    
    # Application Services
    gen.add_application_service(
        "Inference API",
        description="REST API for model inference"
    )
    
    gen.add_application_service(
        "Embedding Service",
        description="Service for generating text embeddings"
    )
    
    # Application Interfaces
    gen.add_application_interface(
        "Chat API",
        description="REST API for chat interactions"
    )
    
    gen.add_application_interface(
        "Admin Console",
        description="Web UI for administration"
    )
    
    return gen


if __name__ == "__main__":
    # Generate sample architecture
    gen = create_sample_ai_architecture()
    
    print("ADOIT Excel Generator - Sample Output")
    print("=" * 50)
    print(gen)
    print("\nElement Summary:")
    for element_type, count in gen.get_summary().items():
        print(f"  {element_type}: {count}")
    
    # Save to file
    output_path = "sample_ai_architecture.xlsx"
    gen.save(output_path, use_template=False)
    print(f"\nSaved to: {output_path}")
